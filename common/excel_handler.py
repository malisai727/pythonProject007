import openpyxl


class ExcelHandler():

    def __init__(self,filename,sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name
    #打开工作簿
    def open(self):
        self.workbook = openpyxl.load_workbook(self.filename)
        self.sheet = self.workbook[self.sheet_name]
    #关闭工作簿
    def close(self):
        self.workbook.save(self.filename)
        self.workbook.close()

    #按行读取储存到列表
    def read_row_list(self):
        self.open()
        datas = list(self.sheet.rows)
        title_list = []
        for data in datas[0]:
            title_list.append(data.value)
        cases_list = []
        for data in datas[1:]:
            cell_list = []
            for cell in data:
                cell_list.append(cell.value)
            cases =  dict(zip(title_list,cell_list))
            cases_list.append(cases)
        self.close()
        return cases_list

    #按行读取储存到类
    def read_row_obj(self):
        self.open()
        cases_obj_list = []
        enumerate
        class CasesObj():
            def __init__(self,cases):
                for case in cases:
                    setattr(self,case[0],case[1])
        datas = list(self.sheet.rows)
        title_list = []
        for data in datas[0]:
            title_list.append(data.value)
        for data in datas[1:]:
            cell_list = []
            for cell in data:
                cell_list.append(cell.value)
            cases = list(zip(title_list, cell_list))
            cases_obj = CasesObj(cases)
            cases_obj_list.append(cases_obj)
        self.close()
        return cases_obj_list

    #读取指定列储存到类
    def read_designated_column(self,column_list):
        self.open()
        cases_obj_list = []
        class Cases():
            def __init__(self,cases):
                for case in cases:
                    setattr(self,case[0],case[1])
        for row_nmb in range(1,self.sheet.max_row+1):
            if row_nmb == 1:
                title_list = []
                for column_num in column_list:
                    cell_value = self.sheet.cell(row=row_nmb,column=column_num).value
                    title_list.append(cell_value)
            else:
                data_list = []
                for column_num in column_list:
                    cell_value = self.sheet.cell(row=row_nmb,column=column_num).value
                    data_list.append(cell_value)
                cases = list(zip(title_list,data_list))
                cases_obj = Cases(cases)
                cases_obj_list.append(cases_obj)
        self.close()
        return cases_obj_list

    #回写excel
    def write(self,row,column,value):
        self.open()
        self.sheet.cell(row=row,column=column,value=value)
        self.close()












if __name__ == '__main__':
    excel_handler = ExcelHandler('../data/cases1.xlsx', 'login')
    a = excel_handler.read_row_obj()
    for i in a:
        print(i.case_id)

